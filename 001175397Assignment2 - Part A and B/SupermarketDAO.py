# SupermarketDAO.py
from openpyxl.chart import BarChart, Reference
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Alignment
from Product import Product
from Transaction import Transaction

class SupermarketDAO:
    def __init__(self, database_name):
        self.database_name = database_name
        self.conn = sqlite3.connect(self.database_name)
        self.cursor = self.conn.cursor()

    def __del__(self):
        self.conn.close()

    def addProductToDB(self, product):
        try:
            self.cursor.execute("INSERT INTO Product (Barcode, Name, Description, Price) VALUES (?, ?, ?, ?)",
                               (product.get_barcode(), product.get_name(), product.get_description(), product.get_price()))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False

    def listAllProducts(self):
        try:
            self.cursor.execute("SELECT * FROM Product ORDER BY Barcode ASC")
            products_data = self.cursor.fetchall()

            products = []

            for product_data in products_data:
                barcode, name, description, price = product_data
                product = Product(barcode, name, description, price)
                products.append(product)

            return sorted(products, key=lambda x: x.get_barcode())
        except sqlite3.Error as e:
            print("Error: An error occurred while retrieving products from the database:", e)
            return []

    def findProduct(self, barcode):
        self.cursor.execute("SELECT * FROM Product WHERE Barcode = ?", (barcode,))
        product_data = self.cursor.fetchone()

        if product_data:
            barcode, name, description, price = product_data
            product = Product(barcode, name, description, price)
            return product
        else:
            return None

    def listAllTransactions(self):
        self.cursor.execute("SELECT * FROM Transactions ORDER BY Date ASC")
        transaction_data = self.cursor.fetchall()
        transactions = []

        for data in transaction_data:
            date, barcode, amount = data
            transaction = Transaction(date, barcode, amount)
            transactions.append(transaction)

        return transactions
    
    def addTransactionToDB(self, transaction):
        try:
            self.cursor.execute("INSERT INTO Transactions (Date, Barcode, Amount) VALUES (?, ?, ?)",
                            (transaction.get_date(), transaction.get_barcode(), transaction.get_amount()))
            self.conn.commit()
            return True
        except sqlite3.Error as e:
            print("Error: An error occurred while adding a transaction to the database:", e)
            return False

    def generate_excel_report(self, filename):
        transactions = self.listAllTransactions()

        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers = ["Date", "Barcode", "Amount"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.alignment = Alignment(horizontal='center')

        for row_num, transaction in enumerate(transactions, 2):
            date = transaction.get_date()
            barcode = transaction.get_barcode()
            amount = transaction.get_amount()
            
            ws.cell(row=row_num, column=1, value=date)
            ws.cell(row=row_num, column=2, value=barcode)
            ws.cell(row=row_num, column=3, value=amount)

        wb.save(filename)


    def displayBarchartOfProductsSold(self, filename='ProductSalesChart.xlsx'):
        # Step 1: Query the database and calculate the counts
        self.cursor.execute("SELECT Barcode, SUM(Amount) FROM Transactions GROUP BY Barcode")
        barcode_counts = self.cursor.fetchall()

        # Step 2: Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sales Data'

        # Step 3: Add headers to the worksheet
        ws.append(['Product Name', 'Quantities Sold'])

        # Step 4: Populate the worksheet with product names and sales data
        for barcode, count in barcode_counts:
            self.cursor.execute("SELECT Name FROM Product WHERE Barcode = ?", (barcode,))
            product_data = self.cursor.fetchone()
            if product_data:
                product_name = product_data[0]
                ws.append([product_name, count])

        # Step 5: Create a bar chart
        chart = BarChart()
        chart.type = "col"  # Column chart
        chart.style = 10  # Choose a style
        chart.title = "Product Sales"
        chart.x_axis.title = "Product Name"
        chart.y_axis.title = "Quantities Sold"

        # Calculate the data range for the chart
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Step 6: Add the chart to the worksheet
        ws.add_chart(chart, "D2")  # Place the chart on the worksheet

        # Step 7: Save the workbook
        wb.save(filename)
        print(f"Bar chart saved in '{filename}'")